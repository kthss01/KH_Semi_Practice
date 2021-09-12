# Selenium import
from selenium import webdriver

import time
import copy
import random
import os
from urllib.request import urlretrieve

# USB 관련 에러 처리
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])

# 크롬 드라이버 생성
driver = webdriver.Chrome('./Python/chromedriver.exe', options=options)

# 사이트 접속하기
driver.get('https://www.wadiz.kr/web/wreward/main?keyword=&endYn=ALL&order=recommend')

time.sleep(2)

# 스크롤 내리기
# lazy loading 먹이려면 일일히 내려가야함
from selenium.webdriver.common.keys import Keys

repeat = 8 # 한번에 48개씩 가져옴 
# repeat *= 5 # (48 * 5 = 240) + 48 (더 가져오기 도함)
repeat *= 10 # (48 * 10 = 480) + 48 = 524

for _ in range(repeat):
    driver.find_element_by_tag_name('body').send_keys(Keys.PAGE_DOWN)
    time.sleep(1)
    
## codes
codes = driver.find_elements_by_css_selector('div.CategoryCircleList_list__2YBF3 a')

code_list = []

for code in codes:
    # print(code.text)
    if code.text != '':
        code_list.append(code.text)
    
# code next button 
driver.find_element_by_css_selector('#main-app > div.MainWrapper_content__GZkTa > div > div.RewardCategoryCircleList_container__1GDge > div > button.CategoryCircleList_next__1mHyX').click()
time.sleep(2)

codes = driver.find_elements_by_css_selector('div.CategoryCircleList_list__2YBF3 a')

for code in codes:
    # print(code.text)
    if code.text != '':
        code_list.append(code.text)
    
# print(code_list)

########################################################
proj_no = 1

users = ['103', '104', '108', '109', '114', '141']

charge = [0, 2000, 2500, 3000, 4000]

projects = []
project_dict = {
    'proj_no' : '',
    'url': '',
    'thumbnail_link' : '',
    'user_no' : '',
    'name' : '',
    'goal' : '',
    'present' : '',
    'date' : '',
    'charge' : '',
    'support' : '',
    'summary' : '',
    'category_no' : '',
    'file_name' : '', # 추후에 디비에서 file_name => file_no로 변경
}

# 프로젝트 처리

items = driver.find_elements_by_css_selector('div.ProjectCardList_item__1owJa')

for item in items:
    project = copy.deepcopy(project_dict)
    
    # 번호 증가
    project['proj_no'] = proj_no
    proj_no += 1
    
    # 유저번호 랜덤하게 선택
    project['user_no'] = random.choice(users)
    
    # 배송료 랜덤하게 선택
    project['charge'] = random.choice(charge)
    
    # 먼저 각 페이지 들어가기전 정보부터 스크래핑
    # thumbnail
    thumbnail = item.find_element_by_css_selector('span.CommonCard_background__3toTR').get_attribute("style")
    thumbnail = thumbnail[thumbnail.find("url(")+5:-3]
    project['thumbnail_link'] = thumbnail
    
    # project_name
    name = item.find_element_by_css_selector('p.CommonCard_title__1oKJY').text
    project['name'] = name
    
    # category_name => no
    category = item.find_element_by_css_selector('span.RewardProjectCard_category__2muXk').text
    project['category_no'] = code_list.index(category) + 1 # code_list에서 찾아서 + 1 0이 시작이라서
    
    # detail 페이지 url
    url = item.find_element_by_css_selector('a.CardLink_link__1k83H').get_attribute('href')
    project['url'] = url
    
    print(project)
    
    projects.append(project)

for project in projects:
    # detail page로 이동
    url = project['url']
    driver.get(url)
    
    time.sleep(2)
    
    # goal date 처리
    goal_date = driver.find_element_by_css_selector('#container > div.reward-body-wrap > div > div.wd-ui-info-wrap > div.wd-ui-sub-campaign-info-container > div > div > section > div.wd-ui-campaign-content > div > div:nth-child(6) > div > p:nth-child(1)').text
    # print(goal_date)
    goal, date = goal_date.split('\n')
    goal = goal[5:-1]
    date = date[date.find('-')+1:]
    project['goal'] = goal
    project['date'] = date
    
    # present
    present = driver.find_element_by_css_selector('#container > div.reward-body-wrap > div > div.wd-ui-info-wrap > div.wd-ui-sub-opener-info > div.project-state-info > div.state-box > p.total-amount > strong').text
    project['present'] = present
    
    # support
    support = driver.find_element_by_css_selector('#container > div.reward-body-wrap > div > div.wd-ui-info-wrap > div.wd-ui-sub-opener-info > div.project-state-info > div.state-box > p.total-supporter > strong').text
    project['support'] = support
    
    # summary
    summary = driver.find_element_by_css_selector('#container > div.reward-body-wrap > div > div.wd-ui-info-wrap > div.wd-ui-sub-campaign-info-container > div > div > section > div.campaign-summary').text
    project['summary'] = summary
    
    # 이미지 다운
    img_folder = './project_img'
    if not os.path.isdir(img_folder): # 폴더 없으면 폴더 생성
        os.mkdir(img_folder)

    thumbnail = project['thumbnail_link']
    urlretrieve(thumbnail, f'{img_folder}/{project["proj_no"]}.jpg') 
    
    project['file_name'] = f'{project["proj_no"]}.jpg'
    
    print(project)

driver.close()

# 엑셀 처리
from openpyxl import Workbook

wb = Workbook()

ws_code = wb.create_sheet('PROJECT_CODE')
ws_project = wb.create_sheet('PROJECT')

# codes 처리
header = ['CATEGORY_NO', 'CATEGORY_NAME']

ws_code.append(header)
category_no = 1

for code in code_list:
    ws_code.append([category_no, code])
    category_no += 1
    
# project 처리

header = ['PROJECT_CODE', 'USER_NO', 'PROJECT_NAME', 'AMOUNT_GOAL', 
          'AMOUNT_PRESENT', 'DDLN', 'DELIVERY_CHARGE', 'SUPPORT_NUM', 'DETAIL_INTRO',
          'CATEGORY_NO', 'FILE_NO']

ws_project.append(header)

for p in projects:
    del(p['url'])
    del(p['thumbnail_link'])
    ws_project.append(list(p.values()))
    
wb.save('project.xlsx')