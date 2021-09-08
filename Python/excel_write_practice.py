from openpyxl import Workbook

# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트 생성
write_ws = write_wb.create_sheet('시트생성')

# Sheet1에 입력
write_ws = write_wb.active
write_ws['A1'] = '숫자'

# 행 단위로 추가
write_ws.append([1,2,3])

# 셀 단위로 추가
write_ws.cell(5, 5, '5행 5열')

write_wb.save('test.xlsx')