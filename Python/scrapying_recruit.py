# Selenium import
from selenium import webdriver

import time
import copy

# USB 관련 에러 처리
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])

# 크롬 드라이버 생성
driver = webdriver.Chrome('./Python/chromedriver.exe', options=options)

# 사이트 접속하기
driver.get('http://job2.wadiz.kr/recruit/info2')

time.sleep(2)

# 직무 코드 scraping
codes = driver.find_elements_by_css_selector('a.fake-menu')

##############################

recruit_list = []
recruit_dict = {
    'title' : '공고명',
    'code' : '직무구분',
    
    'date' : '공고기간',
    'start' : '공고시작일',
    'end' : '공고종료일',
    
    'time' : '공고종류',
    
    'content1' : '공고소개',
    'content2' : '주요업무',
    'content3' : '자격요건',
    'content4' : '우대사항',
    'content5' : '혜택및복지',
    'content6' : '기타사항',
}

# 공고 처리

# 공고 페이지네이션 접근
paging = driver.find_elements_by_css_selector('div.paging a.page-number')

plen = len(paging)
#plen = 1 # test
# 매 클릭마다 찾아야 문제 안생김
for i in range(plen):
    paging = driver.find_elements_by_css_selector('div.paging a.page-number')
    paging[i].click()
    time.sleep(2)
    
    # 각 공고 접근
    bbs_item = driver.find_elements_by_css_selector('li.bbs-item')
    bbslen = len(bbs_item)
    #bbslen = 1 # test
    for j in range(bbslen):
        recruit = copy.deepcopy(recruit_dict)
        
        recruit['code'] = bbs_item[j].find_element_by_css_selector('span.col-sub-title').text
        recruit['title'] = bbs_item[j].find_element_by_css_selector('span.col-title').text
        recruit['time'] = bbs_item[j].find_element_by_css_selector('a.col-flag').text
        recruit['date'] = bbs_item[j].find_element_by_css_selector('span.col-span').text
        
        recruit['code'] = recruit['code'][1:-1]
        
        start, end = recruit['date'].split('~')
        recruit['start'] = start.strip()
        recruit['end'] = end.strip()
        
        print(recruit['title'])
        
        bbs = bbs_item[j].find_element_by_css_selector('a')
        bbs.click()
        time.sleep(2)
        
        # 구조가 너무 달라서 넘김 수작업 가자
        if '상시 인재 Pool 등록' in recruit['title']:
            driver.back() # 뒤로가기
            continue
        if '심사/심의' in recruit['title']:
            driver.back() # 뒤로가기
            continue
        if '펀딩 영업 PD' in recruit['title']:
            driver.back() # 뒤로가기
            continue
        
        # 각 공고 페이지 내용 접근
        contents = driver.find_elements_by_css_selector('div.detail-content p')
        
        temp = ''
        for content in contents:
            temp += content.text

        # 문자열 처리
        # print(temp)
        if len(temp.split('[')) == 5:
            c1, c2, c3, c4, c5 = temp.split('[')
            c6 = c5[c5.find('**기타 사항')+len('**기타 사항'):]
            c5 = c5[c5.find(']')+1:c5.find('**')-1]
        else:
            c1, c2, c3, c4, c5, c6 = temp.split('[')
            c5 = c5[c5.find(']')+1:]
            c6 = c6[c6.find(']')+1:]
            
        c1 = c1[c1.find(']')+1:]
        c2 = c2[c2.find(']')+1:]
        c3 = c3[c3.find(']')+1:]
        c4 = c4[c4.find(']')+1:]
        # print(c1)
        # print(c2)
        # print(c3)
        # print(c4)
        # print(c5)
        # print(c6)

        recruit['content1'] = c1.strip()
        recruit['content2'] = c2.strip()
        recruit['content3'] = c3.strip()
        recruit['content4'] = c4.strip()
        recruit['content5'] = c5.strip()
        recruit['content6'] = c6.strip()
                
        print(recruit)
        recruit_list.append(recruit)
        
        driver.back() # 뒤로가기
        time.sleep(2)
        
################################

# 엑셀 처리
from openpyxl import Workbook

wb = Workbook()

# wb.remove_sheet(wb['Sheet']) # 기본 시트 자리인거 같음

ws_code = wb.create_sheet('RECRUIT_CODE')
ws_recruit = wb.create_sheet('RECRUITMENT')

# codes 처리
header = ['직무구분']

row = 1
col = 1

ws_code.cell(row, col, header[0])

for code in codes:
    row += 1
    ws_code.cell(row, col, code.text[1:code.text.find('(')])

# recruit 처리
header = ['공고명', '직무구분', '공고시작일', '공고종료일', '공고소개', '주요업무', '자격요건', '우대사항', '혜택및복지', '기타사항']

ws_recruit.append(header)

for r in recruit_list:
    rList = [ r['title'], r['code'], r['start'], r['end'], r['content1'], r['content2'], r['content3'], r['content4'], r['content5'], r['content6'] ]
    ws_recruit.append(rList)

wb.save('recruit.xlsx')

driver.close()